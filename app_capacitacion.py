# -*- coding: utf-8 -*-
"""
Reporte Profesional de Capacitaciones TALMA
"""

import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# --------------------------------------------------
# CONFIGURACIÓN
# --------------------------------------------------

st.set_page_config(
    page_title="Reporte Capacitaciones TALMA",
    page_icon="📊",
    layout="wide"
)

# --------------------------------------------------
# CSS CORPORATIVO CON MARCO AZUL
# --------------------------------------------------

st.markdown("""
<style>
.main{
    background-color:#F4F7FB;
    padding:10px;
    border:4px solid #004C97;
    border-radius:15px;
}

h1,h2,h3{
    color:#004C97;
}

.stButton>button{
    background-color:#A7D129;
    color:#004C97;
    font-weight:bold;
    border-radius:8px;
}

.stDownloadButton>button{
    background-color:#A7D129;
    color:#004C97;
    font-weight:bold;
    border-radius:8px;
}

.metric-box{
    text-align:center;
    padding:15px;
    border-radius:12px;
    font-weight:bold;
    font-size:18px;
}
.verde{background-color:#28a745;color:white;}
.amarillo{background-color:#ffc107;color:black;}
.rojo{background-color:#dc3545;color:white;}
</style>
""", unsafe_allow_html=True)

# --------------------------------------------------
# HEADER
# --------------------------------------------------

col1, col2 = st.columns([1,6])

with col1:
    st.image("logo.jpg", width=120)

with col2:
    st.markdown("""
    <div style="
    background: linear-gradient(90deg,#004C97,#005EB8);
    padding:20px;
    border-radius:12px;">
    
    <h2 style="color:white;margin:0;">
    📊 Reporte Capacitaciones TALMA
    </h2>
    
    <p style="color:white;margin:0;">
    Control de vigencia de capacitaciones
    </p>
    
    </div>
    """, unsafe_allow_html=True)

st.write("")

# --------------------------------------------------
# VISTA PREVIA DEL FORMATO
# --------------------------------------------------

st.markdown("## 📄 Formato requerido del archivo")
st.info("El archivo Excel debe tener exactamente esta estructura y debe contener la pestaña llamada: Portal Acumulado")

tabla_html = """
<table style="border-collapse:collapse;width:100%;font-size:14px">
<tr style="background:#f0f0f0;text-align:center;font-weight:bold">
<th></th><th></th><th></th><th></th><th></th><th></th><th></th>
<th colspan="5">Curso1</th>
<th colspan="5">Curso2</th>
</tr>
<tr style="background:#e8e8e8;text-align:center;font-weight:bold">
<th>DNI</th><th>NOMBRE COMPLETO</th><th>CARGO</th><th>F. DE INGRESO</th><th>OFICINA</th><th>CENTRO COSTO</th><th>CENTRO COSTO CODIGO</th>
<th>F. DICTADO</th><th>NOTA</th><th>VENCIMIENTO</th><th>VENC. DIAS</th><th>ESTADO</th>
<th>F. DICTADO</th><th>NOTA</th><th>VENCIMIENTO</th><th>VENC. DIAS</th><th>ESTADO</th>
</tr>
<tr style="text-align:center">
<td>---</td><td>---</td><td>---</td><td>dd/mm/yyyy</td><td>---</td><td>---</td><td>---</td>
<td>dd/mm/yyyy</td><td>---</td><td>dd/mm/yyyy</td><td>---</td><td>---</td>
<td>dd/mm/yyyy</td><td>---</td><td>dd/mm/yyyy</td><td>---</td><td>---</td>
</tr>
</table>
"""

st.markdown(tabla_html, unsafe_allow_html=True)
st.markdown("---")

# --------------------------------------------------
# SUBIR ARCHIVO
# --------------------------------------------------

uploaded_file = st.file_uploader("## 📂 Cargar archivo de capacitaciones", type=["xlsx","xlsm"])

# --------------------------------------------------
# PROCESAR ARCHIVO
# --------------------------------------------------

if uploaded_file:

    st.success("Archivo cargado correctamente")

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # VALIDAR NOMBRE DE HOJA
    if "Acumulado Portal" not in wb.sheetnames:
        st.error("❌ El archivo debe contener una hoja llamada 'Acumulado Portal'")
        st.stop()

    ws = wb["Acumulado Portal"]

    ult_fila = ws.max_row
    ult_col = ws.max_column

    # detectar cursos
    cursos = []
    for j in range(8, ult_col+1, 5):
        cursos.append(ws.cell(row=1,column=j).value)

    headers = ["DNI","Nombre Completo","Cargo","F. Ingreso","Oficina","Centro Costo","Centro Costo Codigo",
               "Curso","F. Dictado","Nota","Vencimiento","Venc. Dias","Estado"]

    data = []
    for i in range(2, ult_fila+1):
        fila = [cell.value for cell in ws[i]]
        if fila[0] is None: 
            continue

        dni = fila[0]
        nombre = fila[1]
        cargo = fila[2]
        f_ingreso = fila[3]
        oficina = fila[4]
        centro_costo = fila[5]
        centro_costo_codigo = fila[6]

        for idx, j in enumerate(range(7, ult_col,5)):
            if j+4 >= len(fila):
                break

            curso = cursos[idx]
            f_dictado = fila[j]
            nota = fila[j+1]
            vencimiento = fila[j+2]
            venc_dias = fila[j+3]
            estado = fila[j+4]

            if any([f_dictado, vencimiento]):
                data.append([dni,nombre,cargo,f_ingreso,oficina,
                             centro_costo,centro_costo_codigo,
                             curso,f_dictado,nota,vencimiento,venc_dias,estado])

    df = pd.DataFrame(data, columns=headers)

    # CALCULAR ESTADOS
    hoy = pd.Timestamp.today().normalize()
    df["F. Dictado"] = pd.to_datetime(df["F. Dictado"],errors="coerce",dayfirst=True)
    df["Vencimiento"] = pd.to_datetime(df["Vencimiento"],errors="coerce",dayfirst=True)
    df["Venc. Dias"] = (df["Vencimiento"] - hoy).dt.days

    df.loc[df["Vencimiento"].isna(),"Estado"] = "VIGENTE"
    df.loc[df["Venc. Dias"] <0,"Estado"] = "VENCIDO"
    df.loc[(df["Venc. Dias"]>=0)&(df["Venc. Dias"]<=30),"Estado"] = "POR VENCER"
    df.loc[df["Venc. Dias"]>30,"Estado"] = "VIGENTE"

    # KPIs
    st.markdown("### 📈 Resumen de Capacitaciones")
    vigentes = (df["Estado"]=="VIGENTE").sum()
    por_vencer = (df["Estado"]=="POR VENCER").sum()
    vencidos = (df["Estado"]=="VENCIDO").sum()

    c1,c2,c3 = st.columns(3)
    c1.markdown(f"<div class='metric-box verde'> Vigentes<br><h2>{vigentes}</h2></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='metric-box amarillo'> Por vencer<br><h2>{por_vencer}</h2></div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='metric-box rojo'> Vencidos<br><h2>{vencidos}</h2></div>", unsafe_allow_html=True)

    # TABLA
    st.markdown("### 📊 Vista previa")
    def color_estado(val):
        val = str(val).upper()
        if val=="VIGENTE":
            return "background-color:#28a745;color:white"
        elif val=="POR VENCER":
            return "background-color:#ffc107;color:black"
        elif val=="VENCIDO":
            return "background-color:#dc3545;color:white"
        return ""

    st.dataframe(df.style.applymap(color_estado), use_container_width=True, height=500)

    # EXPORTAR EXCEL CON AJUSTE DE COLUMNAS Y SPINNER
    with st.spinner('⏳ Generando archivo Excel...'):
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        wb2 = openpyxl.load_workbook(output)
        ws2 = wb2.active

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        fill_map = {"VENCIDO": "FF4C4C", "POR VENCER": "FFF2CC", "VIGENTE": "A7D129"}

        for row in ws2.iter_rows(min_row=2):
            estado = row[12].value
            color = fill_map.get(str(estado).upper(), None)
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Ajustar ancho de columnas automáticamente
        for col in ws2.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws2.column_dimensions[col_letter].width = adjusted_width

        output_final = BytesIO()
        wb2.save(output_final)
        output_final.seek(0)

    st.markdown("### 📥 Descargar reporte")
    st.download_button(
        "⬇️ Descargar archivo",
        data=output_final,
        file_name="Capacitaciones_TALMA_Profesional.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




